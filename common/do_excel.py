# -*- coding:utf-8 _*-
""" 
@author:mongo
@file: do_excel.py 
@version:
@time: 2018/12/17 
@email:3126972006@qq.com
@function： 
"""

import json

import openpyxl

from common.request import Request


class Case:
    """
    测试用例封装类
    """

    def __init__(self):
        self.case_id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:
    """
    excel 操作封装类
    """

    def __init__(self, file_name):  # DoExcel初始化函数
        try:
            self.file_name = file_name
            # 打开一个excel文件，返回一个workbook对象实例，把它定义为DoExcel的属性，以便于在这个类的其他地方使用
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            # 文件未找到异常处理
            print('{0} not found, please check file path'.format(file_name))
            raise e

    def get_cases(self, sheet_name):  # 根据sheet名称，获取在这个sheet里面的所有测试用例数据
        sheet = self.workbook[sheet_name]  # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row  # 获取sheet最大行数
        cases = []  # 定义一个列表，用来存放即将要放进去的测试用例
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case = Case()  # 实例化一个case对象，用来存放测试数据
            case.case_id = sheet.cell(row=r, column=1).value  # 取第r行，第1格的值
            case.title = sheet.cell(row=r, column=2).value  # 取第r行，第4格的值
            case.url = sheet.cell(row=r, column=3).value  # 取第r行，第2格的值
            case.data = sheet.cell(row=r, column=4).value  # 取第r行，第3格的值
            case.method = sheet.cell(row=r, column=5).value  # 取第r行，第5格的值
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行，第6格的值
            cases.append(case)  # 将case放到cases 列表里面

        return cases  # for 循环结束后返回cases列表

    def get_sheet_names(self):  # 获取到work boot里面所有的sheet名称的列表
        return self.workbook.sheetnames

    # 根据sheet name 定位到sheet，然后根据case_id定位到行，取到当前行里面actual这个单元格，
    # 然后给他赋值，最后保存当前workbook
    def write_back_by_case_id(self, sheet_name, case_id, actual, result):
        sheet = self.workbook[sheet_name]  # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row  # 获取sheet最大行数
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case_id_r = sheet.cell(r, 1).value  # 获取第r行，第1列，也就是case_id这一列的值
            if case_id_r == case_id:  # 判断excel里面取到的当前行的case id 是否等于传进来的case id
                sheet.cell(r, 7).value = actual  # 写入传进来的actual到当前行的actual列的单元格
                sheet.cell(r, 8).value = result  # 写入传进来的result到当前行的result列的单元格
                self.workbook.save(filename=self.file_name)
                break


if __name__ == '__main__':
    print('comming')
    #  测试一下DoExcel类
    do_excel = DoExcel('../datas/cases.xlsx')  # 实例化一个DoExcel对象
    sheet_names = do_excel.get_sheet_names()  # 获取到work boot里面所有的sheet名称的列表
    print("sheet 名称列表：", sheet_names)
    case_list = ['login', 'register']  # 定义一个执行测试用例的列表
    for sheet_name in sheet_names:
        if sheet_name in case_list:  # 如果当前的这个sheet_name 不在可执行的case_list里面，就不执行
            cases = do_excel.get_cases(sheet_name)
            print(sheet_name + ' 测试用例个数：', len(cases))
            for case in cases:  # 遍历测试用例列表，每进for一次，就取一个case实例
                print("case信息：", case.__dict__)  # 打印case信息
                data = eval(case.data)  # Excel里面取到data是一个字符串，使用eval函数将字符串转换成字典
                resp = Request(method=case.method, url=case.url, data=data)  # 通过封装的Request类来完成接口的调用
                print('status_code:', resp.get_status_code())  # 打印响应码
                resp_dict = resp.get_json()  # 获取请求响应，字典
                # 通过json.dumps函数将字典转换成格式化后的字符串
                resp_text = json.dumps(resp_dict, ensure_ascii=False, indent=4)
                print('response: ', resp_text)  # 打印响应
                # 判断接口响应是否和Excel里面expected的值是否一致
                if case.expected == resp.get_text():
                    print("result : PASS")
                    do_excel.write_back_by_case_id(sheet_name=sheet_name, case_id=case.case_id, actual=resp.get_text(),
                                                   result='PASS')  # 期望结果与实际结果一致，就写入PASS到result这个单元格
                else:
                    print("result : FAIL")
                    do_excel.write_back_by_case_id(sheet_name=sheet_name, case_id=case.case_id,
                                                   actual=resp.get_text(),
                                                   result='FAIL')  # 期望结果与实际结果一致，就写入FAIL到result这个单元格
